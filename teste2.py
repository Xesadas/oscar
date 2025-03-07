import dash
from dash import dcc, html, dash_table
from dash.dependencies import Input, Output
import openpyxl

app = dash.Dash(__name__)

def read_excel_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_data = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        
        # Ler cabeçalhos
        headers = [cell.value for cell in sheet[1] if cell.value is not None]
        
        # Ler dados
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for header, cell in zip(headers, row):
                row_data[header] = cell.value
            data.append(row_data)
        
        sheets_data[sheet_name] = {
            'headers': headers,
            'data': data
        }
    
    return sheets_data

# Carregar dados do Excel
excel_data = read_excel_data('p.xlsx')

app.layout = html.Div([
    html.H1("Visualizador de Planilhas Excel"),
    
    dcc.Dropdown(
        id='sheet-selector',
        options=[{'label': sheet, 'value': sheet} for sheet in excel_data.keys()],
        value=list(excel_data.keys())[0]
    ),
    
    dash_table.DataTable(
        id='table',
        style_table={'overflowX': 'auto'},
        style_cell={
            'minWidth': '100px', 'width': '150px', 'maxWidth': '300px',
            'whiteSpace': 'normal'
        }
    )
])

@app.callback(
    Output('table', 'columns'),
    Output('table', 'data'),
    Input('sheet-selector', 'value')
)
def update_table(selected_sheet):
    sheet_info = excel_data[selected_sheet]
    columns = [{"name": str(i), "id": str(i)} for i in sheet_info['headers']]
    data = sheet_info['data']
    return columns, data

if __name__ == '__main__':
    app.run_server(debug=True)